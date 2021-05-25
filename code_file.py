import pyttsx3
from plyer import notification
import time
import openpyxl
import inflect
from tkinter import *
from PyQt5.QtCore import QObject, QThread, pyqtSignal
import threading

root = Tk()
root.title('Vocabulary tutor')
root.geometry("700x700")
root.configure(background='gray27')

def submit():
    background = threading.Thread(target=values,args=(my_box1.get(),int(my_box2.get()),))   
    background.daemon = True
    background.start()

def destroy():
    root.destroy()
    quit()
    
read="""Welcome to Vocabulary Tutor,
this app allows you to set the timer.
You will be notified with an english
word and its meaning after every
specified time.Please fill the below
requirements to continue... """

my_label0 = Label(root, text=read, font=("Courier",18,"bold italic"),fg="light cyan",bg="gray27")
my_label0.pack(pady=20)

my_label1 = Label(root, text="Enter your name", font=("Helvetica",18),fg="gold",bg="gray27")
my_label1.pack(pady=20)
    
my_box1 = Entry(root)
my_box1.pack(pady=20)

my_label = Label(root, text="Set the timer (in sec)", font=("Times bold",18),fg = "gold",bg="gray27")
my_label.pack(pady=20)

my_box2 = Entry(root)
my_box2.pack(pady=20)

my_button = Button(root,text="Submit",fg='gray1', bg='PaleVioletRed1' ,activebackground='lawn green',command=submit)
my_button.pack(pady=50)

exit_button= Button(root, text="Exit", command=destroy)
exit_button.pack(pady=20)

p=inflect.engine()

file_location="list.xlsx"
workbook=openpyxl.load_workbook(file_location)
sheet=workbook.active

def notify(title,message):
    notification.notify(title=title,message=message,app_icon="icon3.ico",timeout = 25,)

engine = pyttsx3.init()
voices = engine.getProperty('voices')
engine.setProperty('voice',voices[1].id)
engine.setProperty("rate",135)
engine.setProperty("volume",1.0)

f = True

def threadWork(timer):
    global f
    f = False
    time.sleep(timer)
    f = True

def values(n,t):
    name=n
    timer=t
    value1 = 1
    i=1
    while True:
        global f 

        if f:
            value2=1
            call_obj1=sheet.cell(row=value1,column=value2)
            call_obj2=sheet.cell(row=value1,column=value2+1)
            msg="__WORD__:  "+call_obj1.value+"\n\n__MEANING__:  "+call_obj2.value
            sheet.delete_rows(value1)
            workbook.save('list.xlsx')
            text="hellow  " + name + f" here is your {p.ordinal(i)} word "
            engine.say(text)
            notify("VOCABULARY",msg)
            engine.runAndWait()
            value1=value1+1
            i=i+1
            newThread = threading.Thread(target=threadWork, args=(timer,))
            newThread.daemon = True
            newThread.start()

        newThread.join()
        
        
root.mainloop()
